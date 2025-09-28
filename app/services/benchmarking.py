from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Dict, Iterable, List, Sequence, Tuple
from uuid import UUID

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None  # type: ignore


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


@dataclass(slots=True)
class BenchmarkRecord:
    metric_code: str
    segment: str
    region: str
    value: float


@dataclass(slots=True)
class AggregatedBenchmark:
    metric_code: str
    segment_bucket: str
    region_bucket: str
    count: int
    average_value: float
    min_value: float
    max_value: float


@dataclass(slots=True)
class BenchmarkIngestResult:
    tenant_id: UUID
    batch_id: UUID
    total_rows: int
    discarded_rows: int
    aggregations: List[AggregatedBenchmark]


class BenchmarkRepository:
    def store(self, result: BenchmarkIngestResult) -> None:  # pragma: no cover - interface
        raise NotImplementedError

    def list(self, tenant_id: UUID, batch_id: UUID) -> List[AggregatedBenchmark]:  # pragma: no cover - interface
        raise NotImplementedError


class InMemoryBenchmarkRepository(BenchmarkRepository):
    def __init__(self) -> None:
        self._store: Dict[Tuple[UUID, UUID], List[AggregatedBenchmark]] = {}

    def store(self, result: BenchmarkIngestResult) -> None:
        self._store[(result.tenant_id, result.batch_id)] = result.aggregations

    def list(self, tenant_id: UUID, batch_id: UUID) -> List[AggregatedBenchmark]:
        return self._store.get((tenant_id, batch_id), [])

    def clear(self) -> None:
        self._store.clear()


class BenchmarkingService:
    REQUIRED_COLUMNS = {"metric_code", "segment", "region", "value"}
    MAX_FILE_SIZE_BYTES = 2 * 1024 * 1024  # 2MB local limit

    def __init__(self, repository: BenchmarkRepository | None = None) -> None:
        self.repository = repository or InMemoryBenchmarkRepository()

    def ingest_dataset(self, tenant_id: UUID, batch_id: UUID, *, filename: str, content: bytes) -> BenchmarkIngestResult:
        if len(content) > self.MAX_FILE_SIZE_BYTES:
            raise ValueError("Dataset exceeds maximum size of 2MB")

        rows = self._parse_file(filename, content)
        total_rows = len(rows)
        normalized_records: List[BenchmarkRecord] = []
        discarded_rows = 0

        for row in rows:
            try:
                normalized_records.append(self._normalize_row(row))
            except ValueError:
                discarded_rows += 1

        aggregations = self._aggregate(normalized_records)
        result = BenchmarkIngestResult(
            tenant_id=tenant_id,
            batch_id=batch_id,
            total_rows=total_rows,
            discarded_rows=discarded_rows,
            aggregations=aggregations,
        )
        self.repository.store(result)
        return result

    def list_aggregations(self, tenant_id: UUID, batch_id: UUID) -> List[AggregatedBenchmark]:
        return self.repository.list(tenant_id, batch_id)

    def _parse_file(self, filename: str, content: bytes) -> List[Dict[str, str]]:
        lowered = filename.lower()
        if lowered.endswith(".csv"):
            return self._parse_csv(content)
        if lowered.endswith(".xlsx"):
            return self._parse_excel(content)
        raise ValueError("Unsupported file format. Use CSV or XLSX")

    def _parse_csv(self, content: bytes) -> List[Dict[str, str]]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        reader.fieldnames = [_normalize_header(name) for name in reader.fieldnames or []]
        rows = [self._select_required_columns(row) for row in reader]
        return rows

    def _parse_excel(self, content: bytes) -> List[Dict[str, str]]:
        if load_workbook is None:
            raise ValueError("openpyxl is required to process XLSX files")
        stream = io.BytesIO(content)
        workbook = load_workbook(stream, read_only=True)
        sheet = workbook.active
        headers = [_normalize_header(str(cell.value)) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows: List[Dict[str, str]] = []
        for excel_row in sheet.iter_rows(min_row=2):
            row_dict = {headers[idx]: (str(cell.value) if cell.value is not None else "") for idx, cell in enumerate(excel_row)}
            rows.append(self._select_required_columns(row_dict))
        return rows

    def _select_required_columns(self, row: Dict[str, str]) -> Dict[str, str]:
        selected = {column: row.get(column, "") for column in self.REQUIRED_COLUMNS}
        return selected

    def _normalize_row(self, row: Dict[str, str]) -> BenchmarkRecord:
        metric_code = row["metric_code"].strip().upper()
        segment = row["segment"].strip()
        region = row["region"].strip()
        value_str = row["value"].strip()

        if not metric_code:
            raise ValueError("metric_code required")
        if not segment or not region:
            raise ValueError("segment and region required")

        try:
            value = float(value_str)
        except ValueError as exc:
            raise ValueError("value must be numeric") from exc
        if value < 0:
            raise ValueError("value must be non-negative")

        return BenchmarkRecord(
            metric_code=metric_code,
            segment=self._bucketize_segment(segment),
            region=self._bucketize_region(region),
            value=round(value, 2),
        )

    def _bucketize_segment(self, segment: str) -> str:
        normalized = segment.strip().upper()
        if len(normalized) <= 3:
            return f"{normalized}*"
        return f"{normalized[:3]}*"

    def _bucketize_region(self, region: str) -> str:
        normalized = region.strip().upper()
        if len(normalized) <= 2:
            return f"{normalized}*"
        return f"{normalized[:2]}*"

    def _aggregate(self, records: Iterable[BenchmarkRecord]) -> List[AggregatedBenchmark]:
        buckets: Dict[Tuple[str, str, str], Dict[str, float]] = {}
        for record in records:
            key = (record.metric_code, record.segment, record.region)
            stats = buckets.setdefault(key, {"count": 0, "sum": 0.0, "min": record.value, "max": record.value})
            stats["count"] += 1
            stats["sum"] += record.value
            stats["min"] = min(stats["min"], record.value)
            stats["max"] = max(stats["max"], record.value)

        aggregations: List[AggregatedBenchmark] = []
        for (metric_code, segment_bucket, region_bucket), stats in buckets.items():
            if stats["count"] < 3:  # enforce basic k-anonymity
                continue
            average_value = round(stats["sum"] / stats["count"], 2)
            aggregation = AggregatedBenchmark(
                metric_code=metric_code,
                segment_bucket=segment_bucket,
                region_bucket=region_bucket,
                count=int(stats["count"]),
                average_value=average_value,
                min_value=round(stats["min"], 2),
                max_value=round(stats["max"], 2),
            )
            aggregations.append(aggregation)
        return aggregations


__all__ = [
    "AggregatedBenchmark",
    "BenchmarkIngestResult",
    "BenchmarkRecord",
    "BenchmarkRepository",
    "BenchmarkingService",
    "InMemoryBenchmarkRepository",
]
